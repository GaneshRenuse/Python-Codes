# import openpyxl and tkinter modules
from openpyxl import *
from tkinter import *
from PIL import ImageTk, Image
import random

wb = load_workbook('excel.xlsx')
sheet = wb.active


def my_random(d):
    return random.randrange(10**(d-1), 10**d)


def excel():
    # resize the width of columns in
    # Excel spreadsheet
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 50

    # write given data to an excel spreadsheet
    # at particular location
    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Account Type"
    sheet.cell(row=1, column=3).value = "Age"
    sheet.cell(row=1, column=4).value = "Pan card no"
    sheet.cell(row=1, column=5).value = "Contact Number"
    sheet.cell(row=1, column=6).value = "Email id"
    sheet.cell(row=1, column=7).value = "Gender"
    sheet.cell(row=1, column=7).value = "Account no"


def clear():
    # clear the content of text entry box
    name_field.delete(0, END)
    accntype_field.delete(0, END)
    age_field.delete(0, END)
    pan_no_field.delete(0, END)
    contact_no_field.delete(0, END)
    email_id_field.delete(0, END)
    gender_field.delete(0, END)
    atmpin_field.delete(0, END)


def insert():
    pass


# driver code
if __name__ == "__main__":
    # create a gui window
    root = Tk()

    root.configure(background='light blue')
    root.title("GR Bank")
    root.geometry("600x300")

    n = my_random(7)

    heading = Label(root, text="Account registration", bg="light blue")
    name = Label(root, text="Name", bg="light blue")
    accntype = Label(root, text="Account type", bg="light blue")
    age = Label(root, text="Age", bg="light blue")
    pancno = Label(root, text="Pancard no", bg="light blue")
    contact_no = Label(root, text="Contact No.", bg="light blue")
    email_id = Label(root, text="Email id", bg="light blue")
    gender = Label(root, text="Gender", bg="light blue")
    atmpin = Label(root, text="Set ATM pin", bg="light blue")
    accngen = Label(root, text="Account no generated : ", bg="light blue")
    gen_no = Label(root, text=n, bg="light blue")

    heading.grid(row=0, column=1)
    name.grid(row=1, column=0)
    accntype.grid(row=2, column=0)
    age.grid(row=3, column=0)
    pancno.grid(row=4, column=0)
    contact_no.grid(row=5, column=0)
    email_id.grid(row=6, column=0)
    gender.grid(row=7, column=0)
    atmpin.grid(row=8, column=0)
    accngen.grid(row=9, column=0)
    gen_no.grid(row=9, column=1)

    # create a text entry box
    # for typing the information
    name_field = Entry(root)
    accntype_field = Entry(root)
    age_field = Entry(root)
    pan_no_field = Entry(root)
    contact_no_field = Entry(root)
    email_id_field = Entry(root)
    gender_field = Entry(root)
    atmpin_field = Entry(root)

    name_field.grid(row=1, column=1, ipadx="100")
    accntype_field.grid(row=2, column=1, ipadx="100")
    age_field.grid(row=3, column=1, ipadx="100")
    pan_no_field.grid(row=4, column=1, ipadx="100")
    contact_no_field.grid(row=5, column=1, ipadx="100")
    email_id_field.grid(row=6, column=1, ipadx="100")
    gender_field.grid(row=7, column=1, ipadx="100")
    atmpin_field.grid(row=8, column=1, ipadx="100")

    # create a Submit Button and place into the root window
    submit = Button(root, text="Submit", fg="Black", bg="Red")
    submit.grid(row=10, column=1)

    # start the gui
    root.mainloop()
